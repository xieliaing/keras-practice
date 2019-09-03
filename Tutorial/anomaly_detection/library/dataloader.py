from datetime import datetime
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from pandas_datareader import data

def load_news():
    news_data = pd.read_excel('datasets/news.xlsx')
    news_data = news_data.drop('FactSet', 1)
    news_data = news_data.drop('Browser', 1)
    news_data = news_data.drop('Source', 1)
    news_data['Date'] = pd.to_datetime(news_data['Date/Time'])
    news_data = news_data.drop('Date/Time', 1)
    news_data = news_data.dropna()
    news_data = news_data.set_index('Date')
    return news_data

def load_base_rate():
    df = pd.read_csv("datasets/DFF.csv")
    df = df.replace('.', np.NaN).fillna(method='pad')
    df.DFF = df.DFF.astype(float)
    return df.set_index('DATE')


def load_treasury_yeild():
    df = pd.read_csv("datasets/DGS10.csv")
    df = df.replace('.', np.NaN).fillna(method='pad')
    df.DGS10 = df.DGS10.astype(float)
    return df.set_index('DATE')


def load_nasdaq_100():
    df = pd.read_csv("datasets/NASDAQ100.csv")
    df = df.replace('.', np.NaN).fillna(method='pad')
    df.NASDAQ100 = df.NASDAQ100.astype(float)
    return df.set_index('DATE')


def load_stock():
    start = datetime(1995, 3, 31)
    end = datetime.now()
    return data.get_data_yahoo("MSFT", start, end)


def load_sp500():
    wb = load_workbook(filename='datasets/dailypricehistory.xlsx', read_only=True)
    ws = wb['Sheet1']
    index_data = pd.DataFrame({
        'Date': [r[1][0].value for r in enumerate(ws['A6:A7697'])],
        'SP500': [r[1][0].value for r in enumerate(ws['D6:D7697'])]
    })
    return index_data.set_index('Date')


def load_earnings():
    wb = load_workbook(filename='datasets/FinancialStatementFY17Q1.xlsx', read_only=True)
    ws = wb['Quarterly Income Statements']
    # Q1 SEP 30, Q2 DEC 31, Q3 MAR 31, Q4 JUN 30
    # they probably announce later,not sure how to correct for that
    earnings_data = pd.DataFrame({
        'Date': pd.date_range(start=datetime(1995, 3, 31), periods=87, freq='Q'),
        'Revenue': [r.value for r in ws['D6:CL6'][0]],
        'Gross Margin': [r.value for r in ws['D8:CL8'][0]],
        'Operating Income': [r.value for r in ws['D14:CL14'][0]],
        'Diluted EPS': [r.value for r in ws['D27:CL27'][0]]
    })
    return earnings_data.set_index('Date')


def load_investments():
    investments = pd.read_csv("datasets/investments.csv", parse_dates=[1])
    investments = investments.set_index('Date')
    investments = investments.drop('Calendar Year', 1)
    investments['Investment'] = 1
    return investments[['Investment']]


def load_acquisitions():
    acquisitions = pd.read_csv("datasets/acquisitions.csv", parse_dates=[1])
    acquisitions = acquisitions.set_index('Date')
    acquisitions = acquisitions.drop('Year', 1)
    acquisitions['Acquisition'] = 1
    return acquisitions[['Acquisition']]


def load_sec():
    sec = pd.read_csv("datasets/sec_filings.csv", parse_dates=[0])
    sec = sec.set_index('Date')
    return pd.get_dummies(sec.Filing)


def build_dataset():
    msft_data = load_stock().merge(load_earnings(), how='outer', left_index=True, right_index=True)
    msft_data = msft_data.merge(load_sp500(), how='left', left_index=True, right_index=True)
    # drop columns
    msft_data = msft_data.drop('Open', 1)
    msft_data = msft_data.drop('High', 1)
    msft_data = msft_data.drop('Low', 1)
    msft_data = msft_data.drop('Close', 1)
    msft_data = msft_data.drop('Volume', 1)
    # fill missing values
    msft_data['Adj Close'] = msft_data['Adj Close'].fillna(method='pad')
    msft_data['Diluted EPS'] = msft_data['Diluted EPS'].fillna(method='pad')
    msft_data['Gross Margin'] = msft_data['Gross Margin'].fillna(method='pad')
    msft_data['Operating Income'] = msft_data['Operating Income'].fillna(method='pad')
    msft_data['Revenue'] = msft_data['Revenue'].fillna(method='pad')

    msft_data = msft_data.merge(load_base_rate(), how='outer', left_index=True, right_index=True)
    msft_data['DFF'] = msft_data['DFF'].fillna(method='pad')
    msft_data = msft_data.dropna()

    msft_data = msft_data.merge(load_treasury_yeild(), how='outer', left_index=True, right_index=True)
    msft_data['DGS10'] = msft_data['DGS10'].fillna(method='pad')
    msft_data = msft_data.dropna()

    msft_data = msft_data.merge(load_nasdaq_100(), how='outer', left_index=True, right_index=True)
    msft_data['NASDAQ100'] = msft_data['NASDAQ100'].fillna(method='pad')
    msft_data = msft_data.dropna()

    msft_data = msft_data.merge(load_acquisitions(), how='left', left_index=True, right_index=True)
    msft_data['Acquisition'] = msft_data['Acquisition'].fillna(0.0)

    msft_data = msft_data.merge(load_investments(), how='left', left_index=True, right_index=True)
    msft_data['Investment'] = msft_data['Investment'].fillna(0.0)

    msft_data = msft_data.merge(load_sec(), how='left', left_index=True, right_index=True)
    msft_data = msft_data.fillna(0)
    return msft_data
